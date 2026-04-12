"""
创建多种格式的测试Excel模板
用于测试智能列名识别功能
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

def create_styled_excel(filename: str, sheet_name: str, headers: list, data: list):
    """创建带样式的Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_side = Side(style='thin', color='CCCCCC')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    ws.append(headers)
    for row in data:
        ws.append(row)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            if cell.column > 1:
                cell.alignment = Alignment(horizontal="right")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"✓ 已创建: {filename}")

def create_all_templates():
    """创建所有测试模板"""
    output_dir = os.path.join(os.path.dirname(__file__), 'test_templates')
    os.makedirs(output_dir, exist_ok=True)
    
    sample_data = [
        [2024, 'yearly', '2024-01-01', '2024-12-31', 1000000, 850000, 150000, 
         600000, 500000, 100000, 110500, 85000, 0.13, 400000, 0.25, '否', 
         300000, 60000, 1200, 850, 350],
        [2023, 'yearly', '2023-01-01', '2023-12-31', 900000, 750000, 150000, 
         550000, 450000, 100000, 97500, 75000, 0.13, 350000, 0.25, '否', 
         280000, 55000, 1100, 780, 320],
    ]
    
    print("=" * 80)
    print("创建智能识别测试模板")
    print("=" * 80)
    print()
    
    print("1. 标准中文模板（最完整）")
    create_styled_excel(
        os.path.join(output_dir, '01_标准中文模板.xlsx'),
        '财务数据',
        ['财务年度', '周期类型', '周期开始日期', '周期结束日期', 
         '总收入', '应税销售额', '免税销售额', '总支出', 
         '可抵扣支出', '不可抵扣支出', '进项税额', '销项税额',
         '增值税率', '应纳税所得额', '企业所得税率', '是否小微企业',
         '工资薪金', '专项附加扣除', '发票总数', '进项发票数', '销项发票数'],
        sample_data
    )
    
    print("\n2. 英文完整格式")
    create_styled_excel(
        os.path.join(output_dir, '02_英文完整格式.xlsx'),
        'Financial Data',
        ['Fiscal Year', 'Period Type', 'Start Date', 'End Date',
         'Total Revenue', 'Taxable Sales', 'Tax Free Sales', 'Total Expenses',
         'Deductible Expenses', 'Non Deductible Expenses', 
         'Input Tax', 'Output Tax', 'VAT Rate', 'Taxable Income',
         'Corporate Tax Rate', 'Is Small Enterprise',
         'Total Payroll', 'Special Deductions', 'Total Invoices',
         'Input Invoice Count', 'Output Invoice Count'],
        sample_data
    )
    
    print("\n3. 英文简化格式")
    create_styled_excel(
        os.path.join(output_dir, '03_英文简化格式.xlsx'),
        'Data',
        ['Year', 'Type', 'Revenue', 'Taxable Sales', 'Tax Free Sales', 
         'Expenses', 'Deductible', 'Input Tax', 'Output Tax', 
         'VAT Rate', 'Taxable Income', 'CIT Rate'],
        [
            [2024, 'yearly', 1000000, 850000, 150000, 600000, 500000, 110500, 85000, 0.13, 400000, 0.25],
            [2023, 'yearly', 900000, 750000, 150000, 550000, 450000, 97500, 75000, 0.13, 350000, 0.25],
        ]
    )
    
    print("\n4. 混合中英格式")
    create_styled_excel(
        os.path.join(output_dir, '04_混合中英格式.xlsx'),
        '财务',
        ['年度', '类型', '总收入', '应税销售额', '免税销售', 
         '总支出', '可抵扣费用', '进项税', '销项税', 
         '增值税率', '应税收入', '所得税率'],
        [
            [2024, 'yearly', 1000000, 850000, 150000, 600000, 500000, 110500, 85000, 0.13, 400000, 0.25],
            [2023, 'yearly', 900000, 750000, 150000, 550000, 450000, 97500, 75000, 0.13, 350000, 0.25],
        ]
    )
    
    print("\n5. 用户自定义格式（最灵活）")
    create_styled_excel(
        os.path.join(output_dir, '05_用户自定义格式.xlsx'),
        'My Data',
        ['FY', 'Sales', 'Taxable Rev', 'Tax Free', 'Costs', 
         'Ded Cost', 'VAT In', 'VAT Out', 'Rate'],
        [
            [2024, 1000000, 850000, 150000, 600000, 500000, 110500, 85000, 0.13],
            [2023, 900000, 750000, 150000, 550000, 450000, 97500, 75000, 0.13],
        ]
    )
    
    print("\n6. 带说明的工作表")
    wb = Workbook()
    
    ws_info = wb.active
    ws_info.title = '说明'
    
    info_data = [
        ['智能Excel导入测试模板'],
        [''],
        ['使用方法：'],
        ['1. 在"财务数据"工作表中填写您的财务数据'],
        ['2. 系统会自动识别列名，无需严格匹配'],
        ['3. 必填字段：年度、收入、销售、支出、税额相关字段'],
        ['4. 可选字段：日期、周期类型、工资等'],
        [''],
        ['支持的列名格式：'],
        ['- 中文：总收入、应税销售额、税额等'],
        ['- 英文：Revenue、Sales、Tax、VAT等'],
        ['- 混合格式：收入、Taxable Sales等'],
        [''],
        ['示例数据仅供参考，请替换为您的实际数据'],
    ]
    
    for row in info_data:
        ws_info.append(row)
    
    ws_info.column_dimensions['A'].width = 50
    
    ws_data = wb.create_sheet('财务数据')
    headers = ['财务年度', '总收入', '应税销售额', '免税销售额', '总支出', 
               '可抵扣支出', '进项税额', '销项税额', '增值税率']
    ws_data.append(headers)
    ws_data.append([2024, 1000000, 850000, 150000, 600000, 500000, 110500, 85000, 0.13])
    ws_data.append([2023, 900000, 750000, 150000, 550000, 450000, 97500, 75000, 0.13])
    
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    wb.save(os.path.join(output_dir, '06_带说明模板.xlsx'))
    print(f"✓ 已创建: {output_dir}\\06_带说明模板.xlsx")
    
    print("\n" + "=" * 80)
    print("所有测试模板创建完成！")
    print("=" * 80)
    print(f"\n模板保存位置: {output_dir}")
    print("\n推荐测试顺序:")
    print("1. 06_带说明模板.xlsx - 最简单，适合首次测试")
    print("2. 01_标准中文模板.xlsx - 标准格式，所有字段完整")
    print("3. 02_英文完整格式.xlsx - 英文完整格式")
    print("4. 03_英文简化格式.xlsx - 英文简化格式")
    print("5. 04_混合中英格式.xlsx - 混合中英文")
    print("6. 05_用户自定义格式.xlsx - 最灵活，只有9列")
    print()

if __name__ == '__main__':
    create_all_templates()
