# app/parsers/structured_excel_parser.py
import asyncio
from typing import List
from .base_parser import FileParserStrategy


class StructuredExcelParser(FileParserStrategy):
    """
    结构化 Excel 解析器
    
    核心功能:
    1. 读取 Excel 工作表
    2. 提取单元格内容
    3. 保持表格结构
    4. 处理多个工作表
    """
    
    def get_supported_mime_types(self) -> List[str]:
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
            "application/vnd.ms-excel"  # .xls
        ]
    
    async def parse(self, file_bytes: bytes) -> str:
        """
        解析 Excel 文件，提取结构化内容
        
        Args:
            file_bytes: Excel文件的字节流
            
        Returns:
            str: 结构化的文本内容
        """
        if not self.validate_file(file_bytes):
            raise ValueError("Excel文件为空或无效")
        
        try:
            structured_content = await asyncio.to_thread(
                self._extract_structured_content, file_bytes
            )
            
            if not structured_content.strip():
                raise ValueError("Excel文件内容为空")
            
            return structured_content.strip()
            
        except Exception as e:
            raise Exception(f"Excel解析失败: {str(e)}")
    
    def _extract_structured_content(self, file_bytes: bytes) -> str:
        """同步结构化内容提取（在线程池中运行）"""
        import sys
        import io
        
        print(f"[ExcelParser] 开始解析 Excel 文件，大小: {len(file_bytes)} bytes", file=sys.stderr)
        
        try:
            import openpyxl
            from openpyxl.utils.exceptions import InvalidFileException
            
            try:
                workbook = openpyxl.load_workbook(
                    io.BytesIO(file_bytes),
                    data_only=True,
                    read_only=True
                )
            except InvalidFileException as e:
                print(f"[ExcelParser] 无效的 Excel 文件: {e}", file=sys.stderr)
                raise ValueError(f"无效的 Excel 文件格式: {str(e)}")
            
            sheets_content = []
            
            for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                worksheet = workbook[sheet_name]
                
                sheet_title = worksheet.title or f"Sheet{sheet_index + 1}"
                sheet_lines = [f"## 工作表: {sheet_title}\n"]
                
                rows_data = []
                for row in worksheet.iter_rows(max_row=None, max_col=None, values_only=True):
                    row_data = []
                    for cell_value in row:
                        if cell_value is not None:
                            cell_str = str(cell_value).strip()
                            if cell_str:
                                row_data.append(cell_str)
                    
                    if row_data:
                        rows_data.append(" | ".join(row_data))
                
                if rows_data:
                    sheet_lines.append("### 表格内容:")
                    sheet_lines.append("| " + " | ".join(["列"] * len(rows_data[0].split(" | "))) + " |")
                    sheet_lines.append("| " + " | ".join(["---"] * len(rows_data[0].split(" | "))) + " |")
                    for row in rows_data[:100]:
                        sheet_lines.append(f"| {row} |")
                    
                    if len(rows_data) > 100:
                        sheet_lines.append(f"\n*...共 {len(rows_data)} 行，仅显示前100行...*")
                    
                    sheets_content.append("\n".join(sheet_lines))
                else:
                    sheet_lines.append("*（此工作表为空）*")
                    sheets_content.append("\n".join(sheet_lines))
            
            workbook.close()
            
            if not sheets_content:
                raise ValueError("Excel文件中没有找到有效数据")
            
            final_content = "\n\n".join(sheets_content)
            print(f"[ExcelParser] 成功解析 {len(sheets_content)} 个工作表", file=sys.stderr)
            
            return final_content
            
        except ImportError:
            print(f"[ExcelParser] openpyxl 库未安装", file=sys.stderr)
            raise Exception("Excel 解析需要安装 openpyxl 库: pip install openpyxl")
        except Exception as e:
            print(f"[ExcelParser] 解析失败: {e}", file=sys.stderr)
            raise
