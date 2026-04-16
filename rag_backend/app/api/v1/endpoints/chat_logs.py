"""
对话日志 API 接口

提供对话日志的查询、统计等功能
支持用户级权限控制和企业管理员权限
"""

import uuid
from typing import Optional
from datetime import datetime
from urllib.parse import quote
from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from app.models.user import User
from app.services.chat_log_service import chat_log_service
from app.utils.time_utils import now_beijing
from app.schemas.chat_log import (
    ChatLogSessionListResponse,
    ChatLogSessionStatistics,
    ChatLogUserStatistics,
    ChatLogTenantStatistics,
    UserActionLogListResponse,
)
from app.api.deps import get_db, get_current_user_from_token

router = APIRouter()


@router.get("/test")
async def test_endpoint():
    """测试端点"""
    return {"status": "ok", "message": "chat_logs API is working"}


@router.get("/sessions", response_model=ChatLogSessionListResponse)
async def get_chat_log_sessions(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    user_id: Optional[str] = Query(default=None, description="指定用户ID（仅管理员可用）"),
    keyword: Optional[str] = Query(default=None, description="搜索关键词"),
    start_date: Optional[datetime] = Query(default=None, description="开始日期"),
    end_date: Optional[datetime] = Query(default=None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取对话会话列表

    - 普通用户：只能查看自己的会话
    - 企业管理员：可以查看整个企业的会话，或指定用户的会话
    """
    import logging
    logger = logging.getLogger(__name__)

    from app.schemas.chat_log import ChatLogSessionItem

    try:
        logger.info(f"Getting sessions for user: {current_user.id}, is_admin: {current_user.is_admin}")

        filter_user_id = None
        if user_id:
            try:
                filter_user_id = str(uuid.UUID(user_id))
            except ValueError:
                raise HTTPException(status_code=400, detail="无效的用户ID格式")

        result = await chat_log_service.get_sessions(
            current_user_id=str(current_user.id),
            page=page,
            page_size=page_size,
            user_id=filter_user_id,
            keyword=keyword,
            start_date=start_date,
            end_date=end_date,
        )

        sessions_list = []
        for s in result.get("sessions", []):
            try:
                sessions_list.append(ChatLogSessionItem(**s))
            except (ValueError, KeyError) as e:
                raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
            except (OSError, IOError) as e:
                raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
            except Exception as e:
                logger.warning(f"Failed to parse session item: {e}, data: {s}")

        return ChatLogSessionListResponse(
            total=result.get("total", 0),
            page=result.get("page", page),
            page_size=result.get("page_size", page_size),
            sessions=sessions_list,
        )
    except HTTPException:
        raise
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error getting sessions: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取会话列表失败: {str(e)}")


@router.get("/sessions/{session_id}/messages")
async def get_chat_log_messages(
    session_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取会话的所有消息详情

    - 普通用户：只能查看自己的会话消息
    - 企业管理员：可以查看整个企业的会话消息
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        messages = await chat_log_service.get_session_messages(
            session_id=session_id,
            current_user_id=str(current_user.id),
        )

        if not messages:
            logger.warning(f"No messages found for session {session_id}, user {current_user.id}")
            raise HTTPException(status_code=404, detail="会话不存在或无权访问")

        return {"session_id": session_id, "messages": messages}
    except HTTPException:
        raise
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error getting session messages: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取会话消息失败: {str(e)}")


@router.get("/sessions/{session_id}/statistics", response_model=ChatLogSessionStatistics)
async def get_chat_log_session_statistics(
    session_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取会话的统计信息

    返回该会话的 token 使用量、消息数量等信息
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        statistics = await chat_log_service.get_session_statistics(
            session_id=session_id,
            current_user_id=str(current_user.id),
        )

        if not statistics:
            logger.warning(f"No statistics found for session {session_id}, user {current_user.id}")
            raise HTTPException(status_code=404, detail="会话不存在或无权访问")

        return ChatLogSessionStatistics(**statistics)
    except HTTPException:
        raise
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error getting session statistics: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取会话统计失败: {str(e)}")


@router.get("/statistics/user/{user_id}", response_model=ChatLogUserStatistics)
async def get_chat_log_user_statistics(
    user_id: str,
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取指定用户的对话统计

    - 所有用户：只能查看自己的统计
    """
    if str(current_user.id) != user_id:
        raise HTTPException(status_code=403, detail="无权访问此用户的数据")

    statistics = await chat_log_service.get_user_statistics(
        user_id=user_id,
        start_date=start_date,
        end_date=end_date,
    )

    return ChatLogUserStatistics(**statistics)


@router.get("/statistics/tenant", response_model=ChatLogTenantStatistics)
async def get_chat_log_tenant_statistics(
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    tenant_id: Optional[str] = Query(default=None, description="指定租户ID（可选）"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取租户的对话统计（仅管理员可用）

    返回整个企业的 token 使用量、会话数量等信息
    支持按租户筛选
    """
    statistics = await chat_log_service.get_tenant_statistics(
        current_user_id=str(current_user.id),
        start_date=start_date,
        end_date=end_date,
        tenant_id=tenant_id,
    )

    if "error" in statistics:
        raise HTTPException(status_code=403, detail=statistics["error"])

    return ChatLogTenantStatistics(**statistics)


@router.get("/tenants")
async def get_managed_tenants(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取管理员管理的所有租户列表

    返回所有可访问的企业租户信息
    """
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="需要管理员权限")

    tenants = await chat_log_service.get_all_managed_tenants(str(current_user.id))
    return {"success": True, "data": tenants}


@router.get("/export/sessions")
async def export_chat_sessions(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=1000, ge=1, le=5000),
    user_id: Optional[str] = Query(default=None, description="指定用户ID（仅管理员可用）"),
    keyword: Optional[str] = Query(default=None, description="搜索关键词"),
    start_date: Optional[datetime] = Query(default=None, description="开始日期"),
    end_date: Optional[datetime] = Query(default=None, description="结束日期"),
    format: str = Query(default="xlsx", description="导出格式: xlsx 或 csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    导出会话列表为 Excel 或 CSV

    - 普通用户：只能导出自己的会话
    - 企业管理员：可以导出整个企业的会话
    """
    import logging
    logger = logging.getLogger(__name__)

    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    def clean_for_excel(value):
        """清理数据以确保与Excel兼容"""
        if value is None:
            return ""
        str_value = str(value)
        str_value = str_value.replace('\x00', '').replace('\x01', '').replace('\x02', '')
        str_value = str_value.replace('\x03', '').replace('\x04', '').replace('\x05', '')
        str_value = str_value.replace('\x06', '').replace('\x07', '').replace('\x08', '')
        str_value = str_value.replace('\x0b', '').replace('\x0c', '')
        return str_value[:32767]

    try:
        filter_user_id = None
        if user_id:
            try:
                filter_user_id = str(uuid.UUID(user_id))
            except ValueError:
                raise HTTPException(status_code=400, detail="无效的用户ID格式")

        logger.info(f"Exporting sessions for user {current_user.id}, page={page}, page_size={page_size}")

        result = await chat_log_service.get_sessions(
            current_user_id=str(current_user.id),
            page=page,
            page_size=page_size,
            user_id=filter_user_id,
            keyword=keyword,
            start_date=start_date,
            end_date=end_date,
        )

        headers = ["会话ID", "用户ID", "用户名", "标题", "消息数", "创建时间", "更新时间", "最后消息", "Prompt Tokens", "Completion Tokens", "总Tokens"]

        sessions = result.get("sessions", [])
        session_count = len(sessions)

        if format == "csv":
            import csv
            from io import StringIO

            output = StringIO()
            writer = csv.writer(output)

            writer.writerow([clean_for_excel(h) for h in headers])

            for session in sessions:
                try:
                    row = [
                        clean_for_excel(session.get("id", "")),
                        clean_for_excel(session.get("user_id", "")),
                        clean_for_excel(session.get("user_name", "")),
                        clean_for_excel(session.get("title", "")),
                        int(session.get("message_count", 0) or 0),
                        clean_for_excel(session.get("created_at", "")),
                        clean_for_excel(session.get("updated_at", "")),
                        clean_for_excel(session.get("last_message_preview", "")),
                        int(session.get("total_prompt_tokens", 0) or 0),
                        int(session.get("total_completion_tokens", 0) or 0),
                        int(session.get("total_tokens", 0) or 0),
                    ]
                    writer.writerow(row)
                except (ValueError, KeyError) as e:
                    raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
                except (OSError, IOError) as e:
                    raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
                except Exception as e:
                    logger.warning(f"Failed to add session {session.get('id', 'unknown')} to CSV: {e}")
                    continue

            output.seek(0)
            output_bytes = BytesIO(output.getvalue().encode('utf-8-sig'))

            user_name = clean_for_export(current_user.full_name or current_user.username or "user")
            filename = f"{user_name}_对话日志_{now_beijing().strftime('%Y%m%d%H%M')}.csv"
            encoded_filename = quote(filename)

            logger.info(f"Successfully exported {session_count} sessions to CSV")

            return StreamingResponse(
                output_bytes,
                media_type="text/csv; charset=utf-8-sig",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
            )
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "对话日志"

            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for session in sessions:
                try:
                    row = [
                        clean_for_excel(session.get("id", "")),
                        clean_for_excel(session.get("user_id", "")),
                        clean_for_excel(session.get("user_name", "")),
                        clean_for_excel(session.get("title", "")),
                        int(session.get("message_count", 0) or 0),
                        clean_for_excel(session.get("created_at", "")),
                        clean_for_excel(session.get("updated_at", "")),
                        clean_for_excel(session.get("last_message_preview", "")),
                        int(session.get("total_prompt_tokens", 0) or 0),
                        int(session.get("total_completion_tokens", 0) or 0),
                        int(session.get("total_tokens", 0) or 0),
                    ]
                    ws.append(row)
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=ws.max_row, column=col).border = thin_border
                except (ValueError, KeyError) as e:
                    raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
                except (OSError, IOError) as e:
                    raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
                except Exception as e:
                    logger.warning(f"Failed to add session {session.get('id', 'unknown')} to Excel: {e}")
                    continue

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            user_name = clean_for_export(current_user.full_name or current_user.username or "user")
            filename = f"{user_name}_对话日志_{now_beijing().strftime('%Y%m%d%H%M')}.xlsx"
            encoded_filename = quote(filename)

            logger.info(f"Successfully exported {session_count} sessions to Excel")

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
            )
    except HTTPException:
        raise
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error exporting sessions: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出会话失败: {str(e)}")


@router.get("/actions", response_model=UserActionLogListResponse)
async def get_user_action_logs(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    level: Optional[str] = Query(default=None, description="日志等级（INFO, WARNING, ERROR, DEBUG）"),
    action_type: Optional[str] = Query(default=None, description="操作类型"),
    start_date: Optional[datetime] = Query(default=None, description="开始日期"),
    end_date: Optional[datetime] = Query(default=None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取用户操作日志列表

    包含登录、登出、API调用等操作的详细记录
    """
    result = await chat_log_service.get_user_action_logs(
        current_user_id=str(current_user.id),
        page=page,
        page_size=page_size,
        start_date=start_date,
        end_date=end_date,
        level=level,
        action_type=action_type,
    )

    return UserActionLogListResponse(**result)


@router.get("/export/actions")
async def export_user_action_logs(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=1000, ge=1, le=5000),
    level: Optional[str] = Query(default=None, description="日志等级（INFO, WARNING, ERROR, DEBUG）"),
    action_type: Optional[str] = Query(default=None, description="操作类型"),
    start_date: Optional[datetime] = Query(default=None, description="开始日期"),
    end_date: Optional[datetime] = Query(default=None, description="结束日期"),
    format: str = Query(default="xlsx", description="导出格式: xlsx 或 csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    导出用户操作日志为 Excel 或 CSV

    - 普通用户：只能导出自己的操作日志
    - 企业管理员：可以导出整个企业的操作日志
    """
    import logging
    logger = logging.getLogger(__name__)

    from fastapi.responses import StreamingResponse
    from io import BytesIO, StringIO

    def clean_for_export(value):
        """清理数据以确保与导出格式兼容"""
        if value is None:
            return ""
        str_value = str(value)
        str_value = str_value.replace('\x00', '').replace('\x01', '').replace('\x02', '')
        str_value = str_value.replace('\x03', '').replace('\x04', '').replace('\x05', '')
        str_value = str_value.replace('\x06', '').replace('\x07', '').replace('\x08', '')
        str_value = str_value.replace('\x0b', '').replace('\x0c', '')
        return str_value[:32767]

    try:
        logger.info(f"Exporting action logs for user {current_user.id}, page={page}, page_size={page_size}")

        result = await chat_log_service.get_user_action_logs(
            current_user_id=str(current_user.id),
            page=page,
            page_size=page_size,
            start_date=start_date,
            end_date=end_date,
            level=level,
            action_type=action_type,
        )

        headers = ["日志ID", "时间", "等级", "操作类型", "用户ID", "用户名", "IP地址", "详情", "额外信息"]
        logs = result.get("logs", [])
        log_count = len(logs)

        if format == "csv":
            import csv

            output = StringIO()
            writer = csv.writer(output)

            writer.writerow([clean_for_export(h) for h in headers])

            for log in logs:
                try:
                    row = [
                        clean_for_export(log.get("id", "")),
                        clean_for_export(log.get("created_at", "")),
                        clean_for_export(log.get("level", "")),
                        clean_for_export(log.get("action_type", "")),
                        clean_for_export(log.get("user_id", "")),
                        clean_for_export(log.get("user_name", "")),
                        clean_for_export(log.get("ip_address", "")),
                        clean_for_export(log.get("details", "")),
                        clean_for_export(log.get("extra_data", "")),
                    ]
                    writer.writerow(row)
                except (ValueError, KeyError) as e:
                    raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
                except (OSError, IOError) as e:
                    raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
                except Exception as e:
                    logger.warning(f"Failed to add log {log.get('id', 'unknown')} to CSV: {e}")
                    continue

            output.seek(0)
            output_bytes = BytesIO(output.getvalue().encode('utf-8-sig'))

            user_name = clean_for_export(current_user.full_name or current_user.username or "user")
            filename = f"{user_name}_操作日志_{now_beijing().strftime('%Y%m%d%H%M')}.csv"
            encoded_filename = quote(filename)

            logger.info(f"Successfully exported {log_count} action logs to CSV")

            return StreamingResponse(
                output_bytes,
                media_type="text/csv; charset=utf-8-sig",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
            )
        else:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "操作日志"

            ws.append(headers)

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for log in logs:
                try:
                    row = [
                        clean_for_export(log.get("id", "")),
                        clean_for_export(log.get("created_at", "")),
                        clean_for_export(log.get("level", "")),
                        clean_for_export(log.get("action_type", "")),
                        clean_for_export(log.get("user_id", "")),
                        clean_for_export(log.get("user_name", "")),
                        clean_for_export(log.get("ip_address", "")),
                        clean_for_export(log.get("details", "")),
                        clean_for_export(log.get("extra_data", "")),
                    ]
                    ws.append(row)
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=ws.max_row, column=col).border = thin_border
                except (ValueError, KeyError) as e:
                    raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
                except (OSError, IOError) as e:
                    raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
                except Exception as e:
                    logger.warning(f"Failed to add log {log.get('id', 'unknown')} to Excel: {e}")
                    continue

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            output = BytesIO()
            wb.save(output)
            output.seek(0)

            user_name = clean_for_export(current_user.full_name or current_user.username or "user")
            filename = f"{user_name}_操作日志_{now_beijing().strftime('%Y%m%d%H%M')}.xlsx"
            encoded_filename = quote(filename)

            logger.info(f"Successfully exported {log_count} action logs to Excel")

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
            )
    except HTTPException:
        raise
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error exporting action logs: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"导出操作日志失败: {str(e)}")


@router.get("/tenants")
async def get_admin_tenants(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取当前管理员管理的所有企业列表

    用于企业筛选功能
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        if not current_user.is_admin:
            raise HTTPException(status_code=403, detail="需要管理员权限")

        tenants = await chat_log_service.get_all_managed_tenants(str(current_user.id))

        return {
            "tenants": tenants,
            "count": len(tenants)
        }
    except HTTPException:
        raise
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error getting admin tenants: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取企业列表失败: {str(e)}")


@router.get("/data-integrity")
async def get_data_integrity_report(
    user_id: Optional[str] = Query(default=None, description="指定用户ID（仅管理员可用）"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取数据完整性报告

    - 普通用户：只能查看自己的数据完整性报告
    - 企业管理员：可以查看指定用户或整个企业的数据完整性报告

    返回信息：
    - total_messages: 消息总数
    - missing_embedding_count: 缺少embedding的消息数
    - missing_embedding_percentage: 缺少embedding的百分比
    - missing_importance_count: 缺少importance的消息数
    - needs_repair: 是否需要修复
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        report = await chat_log_service.get_data_integrity_report(
            current_user_id=str(current_user.id),
            user_id=user_id
        )
        return report
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error getting data integrity report: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取数据完整性报告失败: {str(e)}")


@router.get("/messages-missing-embedding")
async def get_messages_missing_embedding(
    user_id: Optional[str] = Query(default=None, description="指定用户ID（仅管理员可用）"),
    limit: int = Query(default=100, ge=1, le=1000, description="返回数量限制"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user_from_token),
):
    """
    获取缺少 embedding 的消息列表

    - 普通用户：只能查看自己的消息
    - 企业管理员：可以查看指定用户的消息

    用于数据修复流程，可以批量生成缺失的embedding

    返回信息：
    - id: 消息ID
    - session_id: 会话ID
    - role: 角色
    - content: 内容预览（前100字符）
    - created_at: 创建时间
    """
    import logging
    logger = logging.getLogger(__name__)

    try:
        messages = await chat_log_service.get_messages_missing_embedding(
            user_id=user_id,
            limit=limit
        )
        return {
            "messages": messages,
            "count": len(messages),
            "limit": limit
        }
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"Error getting messages missing embedding: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"获取缺失embedding消息失败: {str(e)}")
