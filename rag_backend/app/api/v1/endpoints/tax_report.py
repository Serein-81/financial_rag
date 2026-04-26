from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import update
from typing import Optional, Dict, Any
from pathlib import Path
import logging
import io
import os
import uuid
import asyncio
from datetime import datetime

from app.api.deps import get_current_user, get_db, CurrentUser
from app.models.tax_report import TaxReport
from app.schemas.tax_report import (
    TaxReportResponse,
    TaxReportStatusResponse,
    TaxReportListResponse,
    TaxReportProcessingCallback,
    TaxTypeEnum,
    TaxReportStatusEnum,
    ManualTaxReportCreate,
)
from app.services.tax_report_service import TaxReportService
from app.services.tax_file_validator import tax_file_validator

router = APIRouter(tags=["税务报告"])
logger = logging.getLogger(__name__)


def _read_file_sync(file_path: str) -> bytes:
    """同步读取文件（用于 asyncio.to_thread）"""
    with open(file_path, "rb") as f:
        return f.read()


def _save_file_sync(file_path, content: bytes):
    """同步保存文件（用于 asyncio.to_thread）"""
    path_str = str(file_path)
    
    if os.path.isabs(path_str):
        directory = os.path.dirname(path_str)
        if directory:
            os.makedirs(directory, exist_ok=True)
    else:
        directory = os.path.dirname(path_str)
        if directory:
            os.makedirs(directory, exist_ok=True)
    
    if os.path.exists(path_str):
        try:
            os.remove(path_str)
        except OSError:
            pass
    
    try:
        with open(path_str, "wb") as f:
            f.write(content)
    except PermissionError:
        if os.path.exists(path_str):
            try:
                os.chmod(path_str, 0o666)
            except OSError:
                pass
        try:
            with open(path_str, "wb") as f:
                f.write(content)
        except Exception as retry_error:
            logger.error(f"重试文件写入仍然失败: {path_str}, 错误: {retry_error}")
            raise


def _extract_pdf_text(content: bytes) -> str:
    """提取 PDF 文本层"""
    try:
        try:
            from pypdf import PdfReader
            logger.info("使用 pypdf 库提取 PDF 文本")
        except ImportError:
            import PyPDF2
            PdfReader = PyPDF2.PdfReader
            logger.info("使用 PyPDF2 库提取 PDF 文本")
        
        pdf_file = io.BytesIO(content)
        pdf_reader = PdfReader(pdf_file)
        text_parts = []
        for page in pdf_reader.pages:
            page_text = page.extract_text()
            if page_text and page_text.strip():
                text_parts.append(page_text)
        return "\n".join(text_parts)
    except ImportError:
        logger.warning("pypdf/PyPDF2 未安装，尝试 pdfplumber")
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text_parts = [page.extract_text() for page in pdf.pages if page.extract_text()]
            return "\n".join(text_parts)
        except ImportError:
            logger.error("无法提取 PDF 内容：缺少 pypdf/PyPDF2 或 pdfplumber")
            return ""
        except (ValueError, KeyError, OSError, IOError) as e:
            logger.error(f"pdfplumber 提取 PDF 错误: {str(e)}")
            return ""
    except Exception as e:
        logger.error(f"PDF 文本提取失败: {str(e)}")
        return ""


async def _extract_with_ocr(content: bytes) -> str:
    """使用 OCR 服务提取文本（扫描件/图片）"""
    import time
    start_time = time.time()
    
    try:
        from app.services.ocr_factory import OCRFactory
        
        ocr_factory = OCRFactory()
        available_engines = ocr_factory.available_engines
        
        logger.info(f"[OCR] ========== 开始 OCR 处理 ==========")
        logger.info(f"[OCR] 可用引擎: {available_engines}")
        logger.info(f"[OCR] 文件大小: {len(content)} bytes ({len(content)/1024:.2f} KB)")
        
        if not available_engines:
            logger.warning("[OCR] 没有可用的 OCR 引擎")
            return ""
        
        # 检测是否为 PDF 文件
        is_pdf = content[:4] == b'%PDF'
        logger.info(f"[OCR] 文件类型检测: {'PDF' if is_pdf else '图片'}")
        
        if is_pdf:
            logger.info("[OCR] 步骤 1: 处理 PDF 文件")
            
            # 尝试使用 Unstructured API（最强大的文档解析服务）
            if 'unstructured' in available_engines:
                logger.info("[OCR] 步骤 1.1: 尝试使用 Unstructured API")
                adapter = ocr_factory.get_adapter('unstructured')
                
                if adapter and hasattr(adapter, 'extract_text'):
                    from pathlib import Path
                    import tempfile
                    import os
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                        tmp.write(content)
                        tmp_path = tmp.name
                    
                    logger.info(f"[OCR] 临时文件: {tmp_path}")
                    
                    try:
                        step_start = time.time()
                        result = await adapter.extract_text(tmp_path)
                        step_duration = time.time() - step_start
                        
                        logger.info(f"[OCR] Unstructured API 响应时间: {step_duration:.2f}s")
                        logger.info(f"[OCR] 提取结果长度: {len(result) if result else 0} 字符")
                        
                        if result and len(result.strip()) > 50:
                            logger.info(f"[OCR] ✅ Unstructured API 提取成功: {len(result)} 字符")
                            logger.info(f"[OCR] 内容预览: {result[:200]}...")
                            logger.info(f"[OCR] ========== OCR 处理完成 ==========")
                            return result
                        else:
                            logger.warning(f"[OCR] Unstructured API 提取结果不足（{len(result) if result else 0} 字符），尝试其他方法")
                            
                    finally:
                        try:
                            os.unlink(tmp_path)
                            logger.info(f"[OCR] 临时文件已删除: {tmp_path}")
                        except:
                            pass
                else:
                    logger.warning("[OCR] Unstructured 适配器不支持 extract_text 方法")
            else:
                logger.warning("[OCR] Unstructured API 不可用，引擎列表: " + str(available_engines))
            
            # 如果 Unstructured API 不可用或失败，尝试 PDF 转图片 OCR
            logger.info("[OCR] 步骤 1.2: 尝试将 PDF 转换为图片进行 OCR 识别")
            try:
                import fitz  # PyMuPDF
                
                step_start = time.time()
                pdf_doc = fitz.open(stream=content, filetype="pdf")
                total_pages = len(pdf_doc)
                
                logger.info(f"[OCR] PDF 页面数: {total_pages}")
                
                all_text = []
                
                for page_num in range(total_pages):
                    page = pdf_doc[page_num]
                    logger.info(f"[OCR] 处理页面 {page_num + 1}/{total_pages}")
                    
                    # 渲染页面为图片（2x 缩放提高质量）
                    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                    img_bytes = pix.tobytes("png")
                    logger.info(f"[OCR] 页面 {page_num + 1} 渲染完成，图片大小: {len(img_bytes)} bytes")
                    
                    # 使用第一个可用的 OCR 引擎处理图片
                    ocr_success = False
                    for engine_name in available_engines:
                        if engine_name == 'unstructured':
                            continue  # 已经尝试过
                        
                        adapter = ocr_factory.get_adapter(engine_name)
                        if adapter and hasattr(adapter, 'extract_text_from_image'):
                            try:
                                page_start = time.time()
                                page_text = await adapter.extract_text_from_image(img_bytes)
                                page_duration = time.time() - page_start
                                
                                logger.info(f"[OCR] {engine_name} 处理页面 {page_num + 1} 完成，耗时: {page_duration:.2f}s")
                                logger.info(f"[OCR] {engine_name} 提取文本长度: {len(page_text) if page_text else 0} 字符")
                                
                                if page_text and len(page_text.strip()) > 10:
                                    all_text.append(page_text)
                                    logger.info(f"[OCR] ✅ {engine_name} 页面 {page_num + 1} OCR 成功")
                                    ocr_success = True
                                    break
                                else:
                                    logger.warning(f"[OCR] ⚠️ {engine_name} 页面 {page_num + 1} 提取文本不足")
                                    
                            except Exception as ocr_error:
                                logger.warning(f"[OCR] ❌ {engine_name} 处理页面 {page_num + 1} 失败: {ocr_error}")
                                continue
                    
                    if not ocr_success:
                        logger.warning(f"[OCR] ⚠️ 页面 {page_num + 1} 所有 OCR 引擎都失败")
                
                pdf_doc.close()
                step_duration = time.time() - step_start
                logger.info(f"[OCR] PDF 转图片 OCR 总耗时: {step_duration:.2f}s")
                
                if all_text:
                    final_text = "\n\n".join(all_text)
                    logger.info(f"[OCR] ✅ PDF 转图片 OCR 成功: {len(final_text)} 字符")
                    logger.info(f"[OCR] 内容预览: {final_text[:200]}...")
                    logger.info(f"[OCR] ========== OCR 处理完成 ==========")
                    return final_text
                else:
                    logger.error("[OCR] ❌ PDF 转图片 OCR 失败，未提取到有效文本")
                    
            except ImportError:
                logger.warning("[OCR] PyMuPDF 未安装，无法进行 PDF 转图片 OCR")
            except Exception as pdf_error:
                logger.error(f"[OCR] PDF 转图片失败: {pdf_error}")
                import traceback
                logger.error(f"[OCR] 详细错误: {traceback.format_exc()}")
        
        # 处理图片文件（PNG, JPG 等）
        else:
            logger.info(f"[OCR] 步骤 2: 处理图片文件")
            logger.info(f"[OCR] 首选引擎: {available_engines[0]}")
            
            adapter = ocr_factory.get_adapter(available_engines[0])
            
            if adapter and hasattr(adapter, 'extract_text_from_image'):
                try:
                    step_start = time.time()
                    result = await adapter.extract_text_from_image(content)
                    step_duration = time.time() - step_start
                    
                    logger.info(f"[OCR] 图片 OCR 响应时间: {step_duration:.2f}s")
                    logger.info(f"[OCR] 提取结果长度: {len(result) if result else 0} 字符")
                    
                    if result and len(result.strip()) > 0:
                        logger.info(f"[OCR] ✅ 图片 OCR 成功: {len(result)} 字符")
                        logger.info(f"[OCR] 内容预览: {result[:200]}...")
                        logger.info(f"[OCR] ========== OCR 处理完成 ==========")
                        return result
                    else:
                        logger.warning("[OCR] ⚠️ 图片 OCR 未提取到有效文本")
                        
                except Exception as img_error:
                    logger.error(f"[OCR] ❌ 图片 OCR 失败: {img_error}")
                    import traceback
                    logger.error(f"[OCR] 详细错误: {traceback.format_exc()}")
            else:
                logger.warning(f"[OCR] 引擎 {available_engines[0]} 不支持图片 OCR")
        
        logger.warning("[OCR] ❌ 所有 OCR 引擎都无法提取有效文本")
        logger.info("[OCR] ========== OCR 处理完成（失败）==========")
        return ""
        
    except ImportError as e:
        logger.error(f"[OCR] ❌ OCR 依赖未安装: {e}")
        logger.info("[OCR] ========== OCR 处理完成（失败）==========")
        return ""
    except Exception as e:
        logger.error(f"OCR 提取失败: {str(e)}")
        return ""


async def _extract_file_content(content: bytes, content_type: str, filename: str) -> str:
    """
    从上传的文件中提取文本内容用于验证
    
    感知层增强：
    - 支持 PDF（文本层）
    - 支持 PDF（扫描件 → OCR）
    - 支持 Excel
    - 支持 CSV
    - 支持图片（PNG, JPG → OCR）
    
    Args:
        content: 文件字节内容
        content_type: 文件的 MIME 类型
        filename: 文件名
        
    Returns:
        str: 提取的文本内容
    """
    import time
    start_time = time.time()
    
    logger.info(f"[感知层] ========== 开始文件内容提取 ==========")
    logger.info(f"[感知层] 文件名: {filename}")
    logger.info(f"[感知层] 内容类型: {content_type}")
    logger.info(f"[感知层] 文件大小: {len(content)} bytes ({len(content)/1024:.2f} KB)")
    
    try:
        if content_type == "text/csv" or filename.lower().endswith(".csv"):
            logger.info("[感知层] 处理 CSV 文件")
            text = content.decode("utf-8", errors="ignore")
            logger.info(f"[感知层] CSV 提取成功: {len(text)} 字符")
            logger.info(f"[感知层] ========== 感知层处理完成 ==========")
            return text
        
        elif "pdf" in content_type.lower() or filename.lower().endswith(".pdf"):
            logger.info("[感知层] 处理 PDF 文件")
            
            # 步骤 1: 尝试提取 PDF 文本层
            logger.info("[感知层] 步骤 1: 提取 PDF 文本层")
            step_start = time.time()
            extracted_text = _extract_pdf_text(content)
            step_duration = time.time() - step_start
            
            logger.info(f"[感知层] PDF 文本层提取完成，耗时: {step_duration:.2f}s")
            logger.info(f"[感知层] PDF 文本提取结果: {len(extracted_text)} 字符")
            
            # 检查是否包含发票关键字段
            invoice_keywords = ["发票", "税额", "金额", "税号", "纳税人", "购买方", "销售方", "增值税", "价税合计"]
            has_invoice_content = any(kw in extracted_text for kw in invoice_keywords)
            
            matched_keywords = [kw for kw in invoice_keywords if kw in extracted_text]
            logger.info(f"[感知层] 检测到的发票关键词: {matched_keywords if matched_keywords else '无'}")
            
            if extracted_text and len(extracted_text.strip()) > 1000 and has_invoice_content:
                logger.info(f"[感知层] ✅ 发票关键字段检测到，使用文本层结果")
                logger.info(f"[感知层] 内容预览: {extracted_text[:300]}...")
                logger.info(f"[感知层] ========== 感知层处理完成 ==========")
                return extracted_text
            
            logger.warning(f"[感知层] ⚠️ PDF 文本不足（{len(extracted_text)} 字符 < 1000）或缺少发票关键字段")
            logger.info("[感知层] 步骤 2: 尝试 OCR 识别")
            
            # 步骤 2: 使用 OCR
            step_start = time.time()
            ocr_text = await _extract_with_ocr(content)
            step_duration = time.time() - step_start
            
            logger.info(f"[感知层] OCR 提取完成，耗时: {step_duration:.2f}s")
            
            if ocr_text and len(ocr_text.strip()) > 0:
                logger.info(f"[感知层] ✅ OCR 提取成功: {len(ocr_text)} 字符")
                logger.info(f"[感知层] 内容预览: {ocr_text[:300]}...")
                logger.info(f"[感知层] ========== 感知层处理完成 ==========")
                return ocr_text
            
            logger.error("[感知层] ❌ PDF 文本提取和 OCR 都失败")
            logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
            return ""
        
        elif "excel" in content_type.lower() or "spreadsheet" in content_type.lower() or \
             filename.lower().endswith((".xlsx", ".xls")):
            logger.info("[感知层] 处理 Excel 文件")
            try:
                import pandas as pd
                excel_file = io.BytesIO(content)
                if filename.lower().endswith(".xlsx"):
                    df = pd.read_excel(excel_file, engine="openpyxl", header=None)
                else:
                    df = pd.read_excel(excel_file, engine="xlrd", header=None)
                extracted_text = df.to_string(index=False, header=False)
                logger.info(f"[感知层] Excel 提取成功: {len(df)} 行")
                
                # 检测发票关键词
                invoice_keywords = ["发票", "税额", "价税合计", "购买方", "销售方", "纳税人", "增值税", "税号"]
                matched_keywords = [kw for kw in invoice_keywords if kw in extracted_text]
                
                if matched_keywords:
                    logger.info(f"[感知层] ✅ 检测到发票关键词: {matched_keywords}")
                else:
                    logger.warning(f"[感知层] ⚠️ 未检测到发票关键词，可能是普通Excel文件")
                
                logger.info(f"[感知层] ========== 感知层处理完成 ==========")
                return extracted_text
            except ImportError:
                logger.warning("[感知层] pandas未安装，尝试使用openpyxl直接读取")
                try:
                    from openpyxl import load_workbook
                    excel_file = io.BytesIO(content)
                    wb = load_workbook(excel_file, read_only=True, data_only=True)
                    text_parts = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            row_text = " ".join(str(cell) for cell in row if cell is not None)
                            if row_text.strip():
                                text_parts.append(row_text)
                    extracted_text = "\n".join(text_parts)
                    logger.info(f"[感知层] Excel 提取成功: {len(text_parts)} 行")
                    
                    # 检测发票关键词
                    invoice_keywords = ["发票", "税额", "价税合计", "购买方", "销售方", "纳税人", "增值税", "税号"]
                    matched_keywords = [kw for kw in invoice_keywords if kw in extracted_text]
                    
                    if matched_keywords:
                        logger.info(f"[感知层] ✅ 检测到发票关键词: {matched_keywords}")
                    else:
                        logger.warning(f"[感知层] ⚠️ 未检测到发票关键词，可能是普通Excel文件")
                    
                    logger.info(f"[感知层] ========== 感知层处理完成 ==========")
                    return extracted_text
                except ImportError:
                    logger.error("[感知层] 无法提取Excel内容：缺少pandas或openpyxl库")
                    logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
                    return ""
                except Exception as e:
                    logger.error(f"[感知层] openpyxl提取Excel失败: {str(e)}")
                    logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
                    return ""
        
        elif "image" in content_type.lower() or filename.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff")):
            logger.info("[感知层] 处理图片文件")
            ocr_text = await _extract_with_ocr(content)
            if ocr_text:
                logger.info(f"[感知层] 图片 OCR 提取成功: {len(ocr_text)} 字符")
                logger.info(f"[感知层] ========== 感知层处理完成 ==========")
                return ocr_text
            logger.error("[感知层] 图片 OCR 提取失败")
            logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
            return ""
    
    except (ValueError, KeyError) as e:
        logger.error(f"[感知层] 数据错误: {str(e)}")
        logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
        return ""
    except (OSError, IOError) as e:
        logger.error(f"[感知层] IO错误: {str(e)}")
        logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
        return ""
    except Exception as e:
        logger.error(f"[感知层] 未知错误: {str(e)}")
        logger.info(f"[感知层] ========== 感知层处理完成（失败）==========")
        return ""


async def get_tax_report_service(db: AsyncSession = Depends(get_db)) -> TaxReportService:
    return TaxReportService(db)


UPLOAD_DIR = Path("uploads/tax_reports")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


async def _process_tax_report_async(
    report_id: str,
    file_path: str,
    user_id: str,
    tenant_id: str,
    tax_type: str,
    minio_path: str,
    original_filename: str,
    content_type: str,
):
    """
    后台异步处理税务报告
    
    四层架构：
    1. 感知层：提取文件文本
    2. 认知层：TaxSpecialist 分析发票（独立唤醒）
    3. 控制层：硬性规则审判 + 人工审核触发
    4. （交易层：最终申报时触发）
    """
    import time
    from app.db.session import AsyncSessionLocal
    from app.services.minio_service import minio_service
    from app.services.invoice import (
        InvoiceCognitionService,
        RiskJudgeEngine,
        HumanReviewTrigger,
        TenantRiskConfig
    )
    from sqlalchemy import update
    
    start_time = time.time()
    
    try:
        logger.info(f"🔄 [TaxReport] 开始后台处理报告: {report_id}")
        
        logger.info(f"⏱️ [Background] Step 1: 读取文件内容... ({time.time() - start_time:.2f}s)")
        file_content = await asyncio.to_thread(_read_file_sync, file_path)
        logger.info(f"⏱️ [Background] 文件读取完成: {len(file_content)} bytes ({time.time() - start_time:.2f}s)")
        
        file_type = "pdf" if file_path.endswith(".pdf") else "excel" if file_path.endswith((".xlsx", ".xls")) else "csv"
        
        file_text = ""
        validation_result_dict = None
        
        try:
            logger.info(f"⏱️ [Background] Step 2: 提取文件文本（感知层）... ({time.time() - start_time:.2f}s)")
            file_text = await _extract_file_content(file_content, f"application/{file_type}", os.path.basename(file_path))
            logger.info(f"⏱️ [Background] 文本提取完成: {len(file_text)} 字符 ({time.time() - start_time:.2f}s)")
            
        except Exception as e:
            logger.warning(f"⚠️ [Background] 文件文本提取失败: {str(e)}")
        
        extraction_data = None
        risk_decision_data = None
        
        async with AsyncSessionLocal() as db:
            service = TaxReportService(db)
            
            await db.execute(
                update(TaxReport)
                .where(TaxReport.id == report_id)
                .values(
                    status="processing",
                    extracted_content=file_text[:10000] if file_text else None,
                    updated_at=datetime.utcnow()
                )
            )
            await db.commit()
            
            # Step 2.5: 检测文件类型
            invoice_keywords = ["发票", "税额", "价税合计", "购买方", "销售方", "纳税人", "增值税", "税号"]
            is_invoice_file = False
            if file_text:
                matched_keywords = [kw for kw in invoice_keywords if kw in file_text]
                is_invoice_file = len(matched_keywords) >= 2  # 至少匹配2个关键词才认为是发票
                logger.info(f"⏱️ [Background] Step 2.5: 文件类型检测... ({time.time() - start_time:.2f}s)")
                logger.info(f"   - 匹配到发票关键词: {matched_keywords if matched_keywords else '无'}")
                logger.info(f"   - 判定为发票文件: {is_invoice_file}")
            
            cognition_service = InvoiceCognitionService()
            
            logger.info(f"⏱️ [Background] Step 3: 认知层分析（独立唤醒 TaxSpecialist）... ({time.time() - start_time:.2f}s)")
            try:
                if is_invoice_file:
                    logger.info(f"   - 使用发票识别模式")
                    extraction = await cognition_service.analyze_invoice(
                        invoice_text=file_text or "无法提取文本",
                        tenant_id=tenant_id,
                        user_id=user_id
                    )
                else:
                    logger.info(f"   - 使用通用文档分析模式（未检测到发票关键词）")
                    extraction = await cognition_service.analyze_non_invoice_document(
                        document_text=file_text or "无法提取文本",
                        original_filename=original_filename,
                        tenant_id=tenant_id,
                        user_id=user_id
                    )
                
                extraction_data = extraction.model_dump()
                logger.info(f"✅ [Background] 认知层分析完成，置信度: {extraction.confidence:.2f}")
                
            except Exception as e:
                logger.warning(f"⚠️ [Background] 认知层分析失败: {str(e)}")
                extraction_data = {"confidence": 0.0, "semantic_suspicion": ["认知层分析失败"]}
            
            logger.info(f"⏱️ [Background] Step 4: 控制层硬性规则审判... ({time.time() - start_time:.2f}s)")
            risk_engine = RiskJudgeEngine()
            
            from app.services.invoice.risk_judge_engine import InvoiceLLMExtraction
            extraction_obj = InvoiceLLMExtraction(**extraction_data) if extraction_data else InvoiceLLMExtraction(confidence=0.0)
            
            risk_decision = risk_engine.judge(extraction_obj)
            risk_decision_data = risk_decision.model_dump()
            
            logger.info(f"✅ [Background] 控制层审判完成:")
            logger.info(f"   - 风险等级: {risk_decision.risk_level.value}")
            logger.info(f"   - 决策: {risk_decision.decision}")
            logger.info(f"   - 需要人工审核: {risk_decision.requires_human_review}")
            
            ai_analysis_result = {
                "extraction": extraction_data,
                "risk_decision": risk_decision_data,
                "processed_at": datetime.utcnow().isoformat()
            }
            
            await db.execute(
                update(TaxReport)
                .where(TaxReport.id == report_id)
                .values(
                    confidence_score=str(extraction_data.get("confidence", 0)),
                    risk_level=risk_decision.risk_level.value,
                    processing_result=ai_analysis_result,
                    key_metrics={
                        "amount": extraction_data.get("amount"),
                        "tax_amount": extraction_data.get("tax_amount"),
                        "tax_rate": extraction_data.get("tax_rate"),
                        "invoice_number": extraction_data.get("invoice_number"),
                        "semantic_suspicion": extraction_data.get("semantic_suspicion", [])
                    },
                    updated_at=datetime.utcnow()
                )
            )
            await db.commit()
            
            if risk_decision.requires_human_review:
                logger.info(f"⏱️ [Background] Step 5: 触发人工审核... ({time.time() - start_time:.2f}s)")
                
                human_review_trigger = HumanReviewTrigger(db)
                review_id = await human_review_trigger.create_review_request(
                    report_id=report_id,
                    extraction_data=extraction_data,
                    risk_decision=risk_decision_data,
                    tenant_id=tenant_id,
                    user_id=user_id
                )
                
                if review_id:
                    await db.execute(
                        update(TaxReport)
                        .where(TaxReport.id == report_id)
                        .values(
                            status="pending_review",
                            needs_human_review="true",
                            review_request_id=review_id,
                            updated_at=datetime.utcnow()
                        )
                    )
                    await db.commit()
                    logger.info(f"✅ [Background] 人工审核请求已创建: {review_id}")
                else:
                    logger.warning(f"⚠️ [Background] 人工审核请求创建失败")
            else:
                await db.execute(
                    update(TaxReport)
                    .where(TaxReport.id == report_id)
                    .values(
                        status="completed",
                        updated_at=datetime.utcnow()
                    )
                )
                await db.commit()
                logger.info(f"✅ [Background] 自动通过，状态更新为 completed")
            
            logger.info(f"✅ [Background] 后台处理完成: {report_id}, 总耗时: {time.time() - start_time:.2f}s")
        
        try:
            logger.info(f"⏱️ [Background] Step 6: 上传到MinIO... ({time.time() - start_time:.2f}s)")
            await minio_service.upload_document_async(
                file_bytes=file_content,
                object_name=minio_path,
                content_type=content_type
            )
            logger.info(f"☁️ [Background] MinIO上传完成: {minio_path} ({time.time() - start_time:.2f}s)")
        except Exception as e:
            logger.warning(f"⚠️ [Background] MinIO上传失败: {str(e)}")
        
    except Exception as e:
        logger.error(f"❌ [Background] 后台处理失败: {report_id}, 错误: {str(e)}")
        try:
            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(TaxReport)
                    .where(TaxReport.id == report_id)
                    .values(
                        status="failed",
                        processing_message=str(e),
                        updated_at=datetime.utcnow()
                    )
                )
                await db.commit()
        except Exception:
            pass


async def _process_manual_tax_report_async(
    report_id: str,
    input_data: Dict[str, Any],
    user_id: str,
    tenant_id: str,
):
    """
    后台异步处理手动录入的税务报告
    
    手动录入的数据直接走控制层进行风险判断
    """
    import time
    from app.db.session import AsyncSessionLocal
    from sqlalchemy import update
    from app.services.invoice import (
        RiskJudgeEngine,
        HumanReviewTrigger,
        InvoiceLLMExtraction
    )
    
    start_time = time.time()
    
    try:
        logger.info(f"🔄 [ManualTaxReport] 开始后台处理手动录入报告: {report_id}")
        
        async with AsyncSessionLocal() as db:
            await db.execute(
                update(TaxReport)
                .where(TaxReport.id == report_id)
                .values(
                    status="processing",
                    updated_at=datetime.utcnow()
                )
            )
            await db.commit()
            
            revenue = input_data.get("revenue", 0)
            taxable_sales = input_data.get("taxable_sales", 0)
            input_tax = input_data.get("input_tax", 0)
            output_tax = input_data.get("output_tax", 0)
            total_invoices = input_data.get("total_invoices", 0)
            
            confidence = 1.0 if revenue > 0 else 0.0
            
            extraction_data = {
                "amount": revenue,
                "tax_amount": output_tax - input_tax if (output_tax and input_tax) else 0,
                "tax_rate": input_data.get("vat_rate", 0.13),
                "invoice_number": f"manual_{report_id[:8]}",
                "invoice_date": datetime.utcnow().strftime("%Y-%m-%d"),
                "invoice_type": "手动录入",
                "seller_name": input_data.get("company_name"),
                "confidence": confidence,
                "semantic_suspicion": []
            }
            
            if revenue > 1000000:
                extraction_data["semantic_suspicion"].append("手动录入金额较大，建议关注")
            if total_invoices == 0:
                extraction_data["semantic_suspicion"].append("无关联发票，可能存在数据不完整")
            
            logger.info(f"⏱️ [Manual] Step 1: 构建提取数据... ({time.time() - start_time:.2f}s)")
            logger.info(f"   - 金额: {revenue:,.2f}")
            logger.info(f"   - 置信度: {confidence}")
            
            logger.info(f"⏱️ [Manual] Step 2: 控制层硬性规则审判... ({time.time() - start_time:.2f}s)")
            risk_engine = RiskJudgeEngine()
            
            extraction_obj = InvoiceLLMExtraction(**extraction_data)
            risk_decision = risk_engine.judge(extraction_obj)
            risk_decision_data = risk_decision.model_dump()
            
            logger.info(f"✅ [Manual] 控制层审判完成:")
            logger.info(f"   - 风险等级: {risk_decision.risk_level.value}")
            logger.info(f"   - 决策: {risk_decision.decision}")
            
            ai_analysis_result = {
                "extraction": extraction_data,
                "risk_decision": risk_decision_data,
                "input_data": input_data,
                "processed_at": datetime.utcnow().isoformat(),
                "source": "manual_entry"
            }
            
            await db.execute(
                update(TaxReport)
                .where(TaxReport.id == report_id)
                .values(
                    confidence_score=str(confidence),
                    risk_level=risk_decision.risk_level.value,
                    processing_result=ai_analysis_result,
                    key_metrics=extraction_data,
                    updated_at=datetime.utcnow()
                )
            )
            await db.commit()
            
            if risk_decision.requires_human_review:
                logger.info(f"⏱️ [Manual] Step 3: 触发人工审核... ({time.time() - start_time:.2f}s)")
                
                human_review_trigger = HumanReviewTrigger(db)
                review_id = await human_review_trigger.create_review_request(
                    report_id=report_id,
                    extraction_data=extraction_data,
                    risk_decision=risk_decision_data,
                    tenant_id=tenant_id,
                    user_id=user_id
                )
                
                if review_id:
                    await db.execute(
                        update(TaxReport)
                        .where(TaxReport.id == report_id)
                        .values(
                            status="pending_review",
                            needs_human_review="true",
                            review_request_id=review_id,
                            updated_at=datetime.utcnow()
                        )
                    )
                    await db.commit()
                    logger.info(f"✅ [Manual] 人工审核请求已创建: {review_id}")
                else:
                    logger.warning(f"⚠️ [Manual] 人工审核请求创建失败")
            else:
                await db.execute(
                    update(TaxReport)
                    .where(TaxReport.id == report_id)
                    .values(
                        status="completed",
                        updated_at=datetime.utcnow()
                    )
                )
                await db.commit()
                logger.info(f"✅ [Manual] 自动通过，状态更新为 completed")
            
            logger.info(f"✅ [Manual] 后台处理完成: {report_id}, 总耗时: {time.time() - start_time:.2f}s")
            
    except Exception as e:
        logger.error(f"❌ [Manual] 后台处理失败: {report_id}, 错误: {str(e)}")
        try:
            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(TaxReport)
                    .where(TaxReport.id == report_id)
                    .values(
                        status="failed",
                        processing_message=str(e),
                        updated_at=datetime.utcnow()
                    )
                )
                await db.commit()
        except Exception:
            pass


@router.post("/upload", response_model=TaxReportResponse, status_code=201)
async def upload_tax_report(
    file: UploadFile = File(...),
    tax_type: str = Query(..., description="税种类型: VAT, INCOME, PERSONAL, CONSUMPTION, BEHAVIOR"),
    user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    上传税务报告文件进行自动处理

    - 支持文件类型: PDF, Excel (.xlsx, .xls), CSV
    - 最大文件大小: 50MB
    - 优化：快速保存文件，立即返回，后台异步处理验证和分析
    - 自动检测重复文件
    """
    import time
    start_time = time.time()
    
    logger.info(f"📤 [TaxUpload] 收到上传请求: {file.filename}, 大小: {file.size}")
    
    if file.size and file.size > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过50MB")

    allowed_types = ["application/pdf", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     "application/vnd.ms-excel", "text/csv"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="不支持的文件类型，请上传PDF、Excel或CSV文件")

    try:
        # Step 0: 检测重复文件
        logger.info("⏱️ [TaxUpload] Step 0: 检测重复文件...")
        service = TaxReportService(db)
        duplicate_result = await service.check_duplicate_report(
            tenant_id=user.tenant_id,
            original_filename=file.filename
        )
        
        if duplicate_result:
            logger.warning(f"⚠️ [TaxUpload] 检测到重复文件: {file.filename}")
            return JSONResponse(
                status_code=409,
                content={
                    "success": False,
                    "error_type": "DUPLICATE_FILE",
                    "message": f"发现重复文件！您之前已上传过「{file.filename}」，上传时间为 {duplicate_result.get('created_at', '未知')}。",
                    "details": {
                        "original_filename": duplicate_result.get('original_filename'),
                        "existing_report_id": duplicate_result.get('report_id'),
                        "existing_status": duplicate_result.get('status'),
                        "existing_confidence_score": duplicate_result.get('confidence_score'),
                        "existing_risk_level": duplicate_result.get('risk_level'),
                        "created_at": duplicate_result.get('created_at'),
                        "suggestion": "如需重新分析，请先删除旧报告后再上传"
                    }
                }
            )
        report_id = str(uuid.uuid4())
        
        ext = os.path.splitext(file.filename)[1] if file.filename else ".pdf"
        saved_filename = f"{report_id}{ext}"
        file_path = UPLOAD_DIR / saved_filename
        
        logger.info("⏱️ [TaxUpload] Step 1: 开始读取文件内容...")
        content = await file.read()
        logger.info(f"⏱️ [TaxUpload] 文件读取完成，耗时: {time.time() - start_time:.2f}s")
        
        logger.info("⏱️ [TaxUpload] Step 2: 开始保存文件到磁盘...")
        await asyncio.to_thread(_save_file_sync, file_path, content)
        logger.info(f"⏱️ [TaxUpload] 文件保存完成，耗时: {time.time() - start_time:.2f}s")
        
        file_size = len(content)
        
        file_type = "pdf"
        if "excel" in file.content_type.lower() or "spreadsheet" in file.content_type.lower():
            file_type = "excel"
        elif "csv" in file.content_type.lower():
            file_type = "csv"
        
        minio_path = f"{user.tenant_id}/{user.id}/tax-report/{report_id}/{saved_filename}"
        
        logger.info(f"💾 [TaxUpload] 文件已保存: {file_path}, 大小: {file_size} bytes")
        
        logger.info("⏱️ [TaxUpload] Step 3: 开始创建数据库记录...")
        report = TaxReport(
            id=report_id,
            user_id=user.id,
            tenant_id=user.tenant_id,
            filename=saved_filename,
            original_filename=file.filename,
            file_type=file_type,
            file_size=file_size,
            minio_path=minio_path,
            tax_type=tax_type.upper(),
            status="pending",
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        
        db.add(report)
        
        logger.info("⏱️ [TaxUpload] Step 4: 开始提交数据库事务...")
        try:
            await db.commit()
            logger.info(f"⏱️ [TaxUpload] 数据库提交完成，耗时: {time.time() - start_time:.2f}s")
        except Exception as db_error:
            logger.error(f"❌ [TaxUpload] 数据库提交失败: {str(db_error)}")
            await db.rollback()
            raise HTTPException(status_code=500, detail=f"数据库提交失败: {str(db_error)}")
        
        try:
            await db.refresh(report)
            logger.info(f"⏱️ [TaxUpload] 数据库刷新完成，耗时: {time.time() - start_time:.2f}s")
        except Exception as refresh_error:
            logger.warning(f"⚠️ [TaxUpload] 数据库刷新失败，使用缓存数据: {str(refresh_error)}")
        
        logger.info(f"✅ [TaxUpload] 数据库记录已创建: {report_id}")
        
        logger.info("⏱️ [TaxUpload] Step 5: 创建后台处理任务...")
        asyncio.create_task(
            _process_tax_report_async(
                report_id=report_id,
                file_path=str(file_path),
                user_id=user.id,
                tenant_id=user.tenant_id,
                tax_type=tax_type,
                minio_path=minio_path,
                original_filename=file.filename,
                content_type=file.content_type,
            )
        )
        
        total_time = time.time() - start_time
        logger.info(f"🚀 [TaxUpload] 快速返回: 报告ID={report_id}, 总耗时: {total_time:.2f}s")
        
        from app.services.operation_log_service import operation_logger, OperationType
        operation_logger.log_operation(
            operation_type=OperationType.TAX_REPORT_UPLOAD,
            user_id=str(user.id),
            tenant_id=str(user.tenant_id),
            resource="tax_report",
            details={
                "user_id": str(user.id),
                "tenant_id": str(user.tenant_id),
                "filename": file.filename,
                "file_size": file_size,
                "file_type": file_type,
                "tax_type": tax_type,
                "report_id": report_id,
                "minio_path": minio_path,
                "result": "success",
                "processing_time": f"{total_time:.2f}s",
                "upload_timestamp": datetime.utcnow().isoformat()
            },
            risk_level="low"
        )
        
        response_data = {
            "id": str(report.id),
            "user_id": str(report.user_id),
            "tenant_id": str(report.tenant_id),
            "filename": report.filename,
            "original_filename": report.original_filename,
            "file_type": report.file_type,
            "file_size": report.file_size,
            "file_size_mb": round(report.file_size / (1024 * 1024), 2) if report.file_size else 0.0,
            "tax_type": TaxTypeEnum(report.tax_type.lower()) if report.tax_type else None,
            "status": TaxReportStatusEnum(report.status),
            "created_at": report.created_at,
            "updated_at": report.updated_at,
        }
        
        logger.info(f"📤 [TaxUpload] 响应数据预览: {response_data}")
        
        return TaxReportResponse(**response_data)

    except HTTPException:
        total_time = time.time() - start_time
        from app.services.operation_log_service import operation_logger, OperationType
        operation_logger.log_operation(
            operation_type=OperationType.TAX_REPORT_UPLOAD,
            user_id=str(user.id),
            tenant_id=str(user.tenant_id),
            resource="tax_report",
            details={
                "user_id": str(user.id),
                "tenant_id": str(user.tenant_id),
                "filename": file.filename if file else "unknown",
                "tax_type": tax_type,
                "result": "failed",
                "error_type": "HTTPException",
                "processing_time": f"{total_time:.2f}s",
                "upload_timestamp": datetime.utcnow().isoformat()
            },
            risk_level="low"
        )
        raise
    except Exception as e:
        import traceback
        total_time = time.time() - start_time
        error_details = {
            "error_type": type(e).__name__,
            "error_message": str(e),
            "traceback": traceback.format_exc(),
            "report_data": {
                "report_id": report.id if 'report' in locals() else None,
                "user_id": user.id if 'user' in locals() else None,
                "tenant_id": user.tenant_id if 'user' in locals() else None,
                "tax_type": tax_type if 'tax_type' in locals() else None,
            }
        }
        logger.error(f"税务报告上传失败: {error_details}", exc_info=True)
        
        from app.services.operation_log_service import operation_logger, OperationType
        operation_logger.log_operation(
            operation_type=OperationType.TAX_REPORT_UPLOAD,
            user_id=str(user.id),
            tenant_id=str(user.tenant_id),
            resource="tax_report",
            details={
                "user_id": str(user.id),
                "tenant_id": str(user.tenant_id),
                "filename": file.filename if file else "unknown",
                "tax_type": tax_type,
                "result": "failed",
                "error_type": type(e).__name__,
                "error_message": str(e),
                "processing_time": f"{total_time:.2f}s",
                "upload_timestamp": datetime.utcnow().isoformat()
            },
            risk_level="medium"
        )
        
        if "validation errors" in str(e).lower() or "pydantic" in str(e).lower():
            logger.error("🔍 [TaxUpload] Pydantic 验证错误详情:")
            for key, value in error_details.items():
                logger.error(f"   {key}: {value}")
        
        raise HTTPException(
            status_code=500,
            detail=f"税务报告上传失败: {str(e)}"
        )


@router.get("", response_model=TaxReportListResponse)
async def list_tax_reports(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    status: Optional[str] = Query(None, description="过滤状态: PENDING, PROCESSING, COMPLETED, FAILED, NEEDS_REVIEW"),
    tax_type: Optional[str] = Query(None, description="过滤税种类型"),
    start_date: Optional[str] = Query(None, description="开始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取当前用户的税务报告列表（租户+用户双重隔离）

    - 支持分页
    - 支持按状态、税种类型、日期范围过滤
    - 按创建时间倒序排列
    - 每个用户只能看到自己的税务报告
    """
    reports, total = await service.list_tax_reports(
        tenant_id=user.tenant_id,
        user_id=str(user.id),  # 用户级隔离：每个用户只能看到自己的提交记录
        skip=skip,
        limit=limit,
        status=status,
        tax_type=tax_type,
        start_date=start_date,
        end_date=end_date,
    )

    # 计算分页信息
    page = (skip // limit) + 1 if limit > 0 else 1
    total_pages = (total + limit - 1) // limit if limit > 0 else 1
    
    return TaxReportListResponse(
        items=[TaxReportResponse.model_validate(r) for r in reports],
        total=total,
        page=page,
        page_size=limit,
        total_pages=total_pages,
    )


@router.post("/test-invoice-extraction")
async def test_invoice_extraction(
    invoice_text: str = Query(..., description="发票文本内容"),
    user: CurrentUser = Depends(get_current_user),
):
    """
    测试发票检测和提取（仅供开发调试使用）
    
    - 直接调用认知层分析发票文本
    - 返回大模型提取的信息和风险判断结果
    """
    try:
        from app.services.invoice import InvoiceCognitionService, RiskJudgeEngine, InvoiceLLMExtraction as InvoiceExtraction
        
        logger.info(f"🧪 [TestInvoice] 开始测试发票提取...")
        logger.info(f"   - 文本长度: {len(invoice_text)} 字符")
        
        cognition_service = InvoiceCognitionService()
        extraction = await cognition_service.analyze_invoice(
            invoice_text=invoice_text,
            tenant_id=user.tenant_id,
            user_id=str(user.id)
        )
        
        extraction_data = extraction.model_dump()
        
        logger.info(f"✅ [TestInvoice] 认知层分析完成")
        logger.info(f"   - 置信度: {extraction.confidence:.2f}")
        logger.info(f"   - 金额: {extraction.amount}")
        logger.info(f"   - 发票号码: {extraction.invoice_number}")
        logger.info(f"   - 语义可疑点: {len(extraction.semantic_suspicion)} 个")
        
        risk_engine = RiskJudgeEngine()
        risk_decision = risk_engine.judge(extraction)
        risk_decision_data = risk_decision.model_dump()
        
        logger.info(f"✅ [TestInvoice] 控制层审判完成")
        logger.info(f"   - 风险等级: {risk_decision.risk_level.value}")
        logger.info(f"   - 需要人工审核: {risk_decision.requires_human_review}")
        
        return {
            "success": True,
            "cognition": {
                "confidence": extraction.confidence,
                "amount": extraction.amount,
                "tax_amount": extraction.tax_amount,
                "tax_rate": extraction.tax_rate,
                "invoice_number": extraction.invoice_number,
                "invoice_date": extraction.invoice_date,
                "invoice_type": extraction.invoice_type,
                "seller_name": extraction.seller_name,
                "buyer_name": extraction.buyer_name,
                "semantic_suspicion": extraction.semantic_suspicion,
            },
            "risk": {
                "risk_level": risk_decision.risk_level.value,
                "decision": risk_decision.decision,
                "requires_human_review": risk_decision.requires_human_review,
                "trigger_rules": risk_decision_data.get("trigger_rules", []),
                "trigger_reasons": risk_decision_data.get("trigger_reasons", []),
            }
        }
        
    except Exception as e:
        logger.error(f"❌ [TestInvoice] 测试失败: {str(e)}")
        import traceback
        logger.error(f"   Trace: {traceback.format_exc()}")
        return {
            "success": False,
            "error": str(e),
            "trace": traceback.format_exc()
        }


@router.get("/statistics")
async def get_tax_report_statistics(
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取税务报告统计信息

    - 按状态统计数量
    - 按税种类型统计数量
    - 总处理时长统计
    - 支持用户级隔离：每个用户只能看到自己的统计数据
    """
    stats = await service.get_statistics(user.tenant_id, str(user.id))
    return stats


@router.get("/{report_id}", response_model=TaxReportResponse)
async def get_tax_report(
    report_id: str,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取税务报告详情（租户+用户双重隔离）

    - 返回报告基本信息
    - 返回处理结果
    - 返回税务验证结果（如有）
    - 用户只能访问自己的报告
    """
    report = await service.get_tax_report(report_id, user.tenant_id, str(user.id))
    if not report:
        raise HTTPException(status_code=404, detail="税务报告不存在")
    return TaxReportResponse.model_validate(report)


@router.get("/{report_id}/status", response_model=TaxReportStatusResponse)
async def get_report_status(
    report_id: str,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取税务报告处理状态（租户+用户双重隔离）

    - 用于轮询查询处理进度
    - 返回当前状态和进度信息
    - 用户只能查看自己的报告状态
    """
    status = await service.get_processing_status(report_id, user.tenant_id, str(user.id))
    if not status:
        raise HTTPException(status_code=404, detail="税务报告不存在")
    return TaxReportStatusResponse(**status)


@router.post("/{report_id}/retry")
async def retry_processing(
    report_id: str,
    background_tasks: BackgroundTasks,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    重试失败的税务报告处理（租户+用户双重隔离）

    - 仅在状态为 FAILED 时可用
    - 重新触发后台处理流程
    - 用户只能重试自己的报告
    """
    report = await service.get_tax_report(report_id, user.tenant_id, str(user.id))
    if not report:
        raise HTTPException(status_code=404, detail="税务报告不存在")

    if report.status != "FAILED":
        raise HTTPException(status_code=400, detail="只能重试失败的报告")

    background_tasks.add_task(
        service.process_tax_report_background,
        report_id,
        user.id,
        user.tenant_id,
    )

    return {"message": "重试任务已提交", "report_id": report_id}


@router.delete("/{report_id}")
async def delete_tax_report(
    report_id: str,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    删除税务报告（租户+用户双重隔离）

    - 仅可删除自己上传的报告
    - 删除时会同时删除关联的文件和文档
    - 用户只能删除自己的报告
    """
    success = await service.delete_tax_report(report_id, user.tenant_id, str(user.id))
    if not success:
        raise HTTPException(status_code=404, detail="税务报告不存在或无权删除")
    return {"message": "税务报告已删除", "report_id": report_id}


@router.post("/manual", response_model=dict)
async def create_manual_tax_report(
    request: ManualTaxReportCreate,
    background_tasks: BackgroundTasks,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
    db: AsyncSession = Depends(get_db),
):
    """
    手动录入税务报告
    
    - 管理员可直接录入财务数据创建税务报告
    - 支持关联已有财务数据
    - 自动走四层架构处理流程（低风险自动通过，高风险触发人工审核）
    """
    try:
        input_data = request.input_data.model_dump()
        
        result = await service.create_manual_tax_report(
            user_id=str(user.id),
            tenant_id=user.tenant_id,
            input_data=input_data,
        )
        
        if result.get("success") is False:
            return result
        
        report_id = result.get("id")
        
        background_tasks.add_task(
            _process_manual_tax_report_async,
            report_id=report_id,
            input_data=input_data,
            user_id=str(user.id),
            tenant_id=user.tenant_id,
        )
        
        return {
            "success": True,
            "message": "税务报告录入成功，正在后台处理",
            "data": result,
            "report_id": report_id,
            "processing": True
        }
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")


@router.post("/callback/processing")
async def processing_callback(
    callback: TaxReportProcessingCallback,
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    处理完成回调（内部接口，供Agent系统调用）

    - 更新报告状态
    - 保存处理结果
    - 自动判断是否需要人工审核
    """
    try:
        await service.update_processing_result(
            report_id=callback.report_id,
            status=callback.status,
            processing_result=callback.processing_result,
            tax_validation_result=callback.tax_validation_result,
            needs_human_review=callback.needs_human_review,
            key_metrics=callback.key_metrics,
            issues=callback.issues,
        )
        return {"message": "回调处理成功"}
    except (ValueError, KeyError) as e:
        logger.error(f"处理回调数据错误: {str(e)}", exc_info=True)
        raise HTTPException(status_code=400, detail=f"回调处理数据错误: {str(e)}")
    except (OSError, IOError) as e:
        logger.error(f"处理回调IO错误: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"回调处理IO错误: {str(e)}")
    except (OSError, IOError) as e:
        raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
    except Exception as e:
        logger.error(f"处理回调失败: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail="回调处理失败")


@router.get("/reviews/pending", response_model=TaxReportListResponse)
async def get_pending_tax_reviews(
    skip: int = Query(0, ge=0),
    limit: int = Query(20, ge=1, le=100),
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取待审核的税务报告列表（租户+用户双重隔离）

    - 只返回需要人工审核的报告
    - 支持分页
    - 按创建时间倒序排列
    - 每个用户只能看到自己的待审核报告
    """
    reports, total = await service.list_tax_reports(
        tenant_id=user.tenant_id,
        user_id=str(user.id),  # 用户级隔离
        skip=skip,
        limit=limit,
        status="pending_review",
    )

    # 计算分页信息
    page = (skip // limit) + 1 if limit > 0 else 1
    total_pages = (total + limit - 1) // limit if limit > 0 else 1
    
    return TaxReportListResponse(
        items=[TaxReportResponse.model_validate(r) for r in reports],
        total=total,
        page=page,
        page_size=limit,
        total_pages=total_pages,
    )
