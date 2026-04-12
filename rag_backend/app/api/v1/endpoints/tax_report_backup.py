from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from typing import Optional
import json
import logging
import io
import asyncio
import os
from pathlib import Path
from datetime import datetime
import uuid

from app.api.deps import get_current_user, get_db, CurrentUser, PaginatedParams
from app.models.tax_report import TaxReport, TaxReportDocument
from app.schemas.tax_report import (
    TaxReportCreate,
    TaxReportResponse,
    TaxReportStatusResponse,
    TaxReportListResponse,
    TaxReportProcessingCallback,
)
from app.services.tax_report_service import TaxReportService
from app.services.tax_file_validator import tax_file_validator

router = APIRouter(tags=["税务报告"])
logger = logging.getLogger(__name__)

UPLOAD_DIR = Path("uploads/tax_reports")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


async def _extract_file_content(content: bytes, content_type: str, filename: str) -> str:
    """
    从上传的文件中提取文本内容用于验证
    
    Args:
        content: 文件字节内容
        content_type: 文件的MIME类型
        filename: 文件名
        
    Returns:
        str: 提取的文本内容
    """
    try:
        if content_type == "text/csv" or filename.lower().endswith(".csv"):
            return content.decode("utf-8", errors="ignore")
        
        elif "pdf" in content_type.lower() or filename.lower().endswith(".pdf"):
            try:
                import PyPDF2
                pdf_file = io.BytesIO(content)
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                text_parts = []
                for page in pdf_reader.pages:
                    text_parts.append(page.extract_text())
                return "\n".join(text_parts)
            except ImportError:
                logger.warning("PyPDF2未安装，尝试使用pdfplumber")
                try:
                    import pdfplumber
                    with pdfplumber.open(io.BytesIO(content)) as pdf:
                        text_parts = [page.extract_text() for page in pdf.pages]
                    return "\n".join(text_parts)
                except ImportError:
                    logger.error("无法提取PDF内容：缺少PyPDF2或pdfplumber库")
                    return ""
                except (ValueError, KeyError) as e:
                    logger.error(f"pdfplumber提取PDF数据错误: {str(e)}")
                    return ""
                except (OSError, IOError) as e:
                    logger.error(f"pdfplumber提取PDF IO错误: {str(e)}")
                    return ""
                except (ValueError, KeyError) as e:
                    raise HTTPException(status_code=400, detail=f"数据错误: {str(e)}")
                except (OSError, IOError) as e:
                    raise HTTPException(status_code=500, detail=f"IO错误: {str(e)}")
                except Exception as e:
                    logger.error(f"pdfplumber提取PDF失败: {str(e)}")
                    return ""
        
        elif "excel" in content_type.lower() or "spreadsheet" in content_type.lower() or \
             filename.lower().endswith((".xlsx", ".xls")):
            try:
                import pandas as pd
                excel_file = io.BytesIO(content)
                if filename.lower().endswith(".xlsx"):
                    df = pd.read_excel(excel_file, engine="openpyxl", header=None)
                else:
                    df = pd.read_excel(excel_file, engine="xlrd", header=None)
                return df.to_string(index=False, header=False)
            except ImportError:
                logger.warning("pandas未安装，尝试使用openpyxl直接读取")
                try:
                    from openpyxl import load_workbook
                    excel_file = io.BytesIO(content)
                    wb = load_workbook(excel_file, read_only=True, data_only=True)
                    text_parts = []
                    for ws in wb.worksheets:
                        for row in ws.iter_rows(values_only=True):
                            row_text = " ".join(str(cell) for cell in row if cell is not None)
                            if row_text.strip():
                                text_parts.append(row_text)
                    return "\n".join(text_parts)
                except ImportError:
                    logger.error("无法提取Excel内容：缺少pandas或openpyxl库")
                    return ""
                except (ValueError, KeyError) as e:
                    logger.error(f"Excel数据错误: {str(e)}")
                    return ""
                except (OSError, IOError) as e:
                    logger.error(f"Excel IO错误: {str(e)}")
                    return ""
                except Exception as e:
                    logger.error(f"Excel提取失败: {str(e)}")
                    return ""
        else:
            try:
                return content.decode("utf-8", errors="ignore")
            except Exception:
                return ""
    except Exception as e:
        logger.error(f"文件内容提取失败: {str(e)}")
        return ""


async def get_tax_report_service(db: AsyncSession = Depends(get_db)) -> TaxReportService:
    return TaxReportService(db)


async def _process_tax_report_async(
    service: TaxReportService,
    report_id: str,
    file_path: str,
    user_id: str,
    tenant_id: str,
    tax_type: str
):
    """
    后台异步处理税务报告
    
    在后台执行文件验证、AI分析和数据库更新
    """
    try:
        logger.info(f"🔄 [TaxReport] 开始后台处理报告: {report_id}")
        
        with open(file_path, "rb") as f:
            file_content = f.read()
        
        file_type = "pdf" if file_path.endswith(".pdf") else "excel" if file_path.endswith((".xlsx", ".xls")) else "csv"
        
        try:
            file_text = await _extract_file_content(file_content, f"application/{file_type}", os.path.basename(file_path))
            
            validation_result = await tax_file_validator.validate_with_ocr_fallback(
                file_text, 
                file_bytes=file_content,
                file_type=f"application/{file_type}"
            )
            
            if validation_result and validation_result.is_valid:
                validation_result_dict = {
                    "confidence": validation_result.confidence,
                    "is_valid": validation_result.is_valid,
                    "found_keywords": validation_result.found_keywords,
                    "missing_indicators": validation_result.missing_indicators,
                    "extracted_info": validation_result.extracted_info,
                    "suggestions": validation_result.suggestions,
                }
            else:
                validation_result_dict = None
                
        except Exception as e:
            logger.warning(f"⚠️ 后台验证失败: {str(e)}")
            validation_result_dict = None
        
        from sqlalchemy import update, text
        from app.db.session import AsyncSessionLocal
        
        async with AsyncSessionLocal() as db:
            await db.execute(
                update(TaxReport)
                .where(TaxReport.id == report_id)
                .values(
                    status="processing",
                    validation_result=validation_result_dict,
                    updated_at=datetime.utcnow()
                )
            )
            await db.commit()
        
        await service.process_tax_report_background(report_id, user_id, tenant_id)
        
        logger.info(f"✅ [TaxReport] 后台处理完成: {report_id}")
        
    except Exception as e:
        logger.error(f"❌ [TaxReport] 后台处理失败: {report_id}, 错误: {str(e)}")
        try:
            from sqlalchemy import update
            from app.db.session import AsyncSessionLocal
            
            async with AsyncSessionLocal() as db:
                await db.execute(
                    update(TaxReport)
                    .where(TaxReport.id == report_id)
                    .values(
                        status="failed",
                        error_message=str(e),
                        updated_at=datetime.utcnow()
                    )
                )
                await db.commit()
        except Exception:
            pass


@router.post("/upload", response_model=TaxReportResponse, status_code=201)
async def upload_tax_report(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    tax_type: str = Query(..., description="税种类型: VAT, INCOME, PERSONAL, CONSUMPTION, BEHAVIOR"),
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    上传税务报告文件进行自动处理

    - 支持文件类型: PDF, Excel (.xlsx, .xls), CSV
    - 最大文件大小: 50MB
    - 优化：快速保存文件，立即返回，后台异步处理验证和分析
    """
    logger.info(f"📤 [TaxUpload] 收到上传请求: {file.filename}, 大小: {file.size}")
    
    if file.size and file.size > 50 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件大小不能超过50MB")

    allowed_types = ["application/pdf", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     "application/vnd.ms-excel", "text/csv"]
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="不支持的文件类型，请上传PDF、Excel或CSV文件")

    try:
        report_id = str(uuid.uuid4())
        
        ext = os.path.splitext(file.filename)[1] if file.filename else ".pdf"
        saved_filename = f"{report_id}{ext}"
        file_path = UPLOAD_DIR / saved_filename
        
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
        
        logger.info(f"💾 [TaxUpload] 文件已保存: {file_path}, 大小: {len(content)} bytes")
        
        result = await service.create_tax_report(
            user_id=user.id,
            tenant_id=user.tenant_id,
            file=file,
            tax_type=tax_type,
            file_validation_result=None,
        )
        
        logger.info(f"✅ [TaxUpload] 数据库记录已创建: {result['id']}")
        
        asyncio.create_task(
            _process_tax_report_async(
                service=service,
                report_id=result["id"],
                file_path=str(file_path),
                user_id=user.id,
                tenant_id=user.tenant_id,
                tax_type=tax_type
            )
        )
        
        logger.info(f"🚀 [TaxUpload] 快速返回: 报告ID={result['id']}")
        
        return TaxReportResponse(**result)

    except ValueError as e:
        logger.error(f"❌ [TaxUpload] ValueError: {str(e)}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"❌ [TaxUpload] 未知错误: {str(e)}")
        raise HTTPException(status_code=500, detail=f"上传失败: {str(e)}")


async def list_tax_reports(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(10, ge=1, le=100, description="每页数量"),
    status: Optional[str] = Query(None, description="过滤状态"),
    tax_type: Optional[str] = Query(None, description="过滤税种类型"),
    search: Optional[str] = Query(None, description="搜索关键词"),
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取税务报告列表
    """
    reports, total = await service.list_tax_reports(
        user_id=user.id,
        tenant_id=user.tenant_id,
        page=page,
        page_size=page_size,
        status=status,
        tax_type=tax_type,
        search=search
    )
    
    return TaxReportListResponse(
        reports=reports,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=(total + page_size - 1) // page_size
    )


@router.get("/statistics", response_model=dict)
async def get_tax_report_statistics(
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取税务报告统计信息
    """
    return await service.get_statistics(user.tenant_id)


@router.get("/{report_id}", response_model=TaxReportResponse)
async def get_tax_report(
    report_id: str,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取单个税务报告详情
    """
    report = await service.get_tax_report(report_id, user.tenant_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    return report


@router.get("/{report_id}/status", response_model=TaxReportStatusResponse)
async def get_tax_report_status(
    report_id: str,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    获取税务报告处理状态
    """
    report = await service.get_tax_report(report_id, user.tenant_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    return TaxReportStatusResponse(
        id=report["id"],
        status=report["status"],
        progress=report.get("progress", 0),
        error_message=report.get("error_message")
    )


@router.post("/{report_id}/retry", response_model=TaxReportResponse)
async def retry_tax_report(
    report_id: str,
    background_tasks: BackgroundTasks,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    重试失败的税务报告处理
    """
    report = await service.get_tax_report(report_id, user.tenant_id)
    if not report:
        raise HTTPException(status_code=404, detail="报告不存在")
    
    if report["status"] not in ["failed", "completed"]:
        raise HTTPException(status_code=400, detail="只能重试失败或已完成的任务")
    
    background_tasks.add_task(
        service.process_tax_report_background,
        report_id,
        user.id,
        user.tenant_id,
    )
    
    return await service.get_tax_report(report_id, user.tenant_id)


@router.delete("/{report_id}", response_model=dict)
async def delete_tax_report(
    report_id: str,
    user: CurrentUser = Depends(get_current_user),
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    删除税务报告
    """
    success = await service.delete_tax_report(report_id, user.tenant_id)
    if not success:
        raise HTTPException(status_code=404, detail="报告不存在")
    return {"success": True, "message": "报告已删除"}


@router.post("/callback/{report_id}", response_model=dict)
async def tax_report_callback(
    report_id: str,
    callback_data: TaxReportProcessingCallback,
    service: TaxReportService = Depends(get_tax_report_service),
):
    """
    接收处理回调（供内部服务调用）
    """
    logger.info(f"📬 [TaxReport] 收到回调: report_id={report_id}, status={callback_data.status}")
    
    if callback_data.status == "completed":
        await service.update_report_completed(
            report_id=report_id,
            analysis_result=callback_data.analysis_result,
            suggestions=callback_data.suggestions
        )
    elif callback_data.status == "failed":
        await service.update_report_failed(
            report_id=report_id,
            error_message=callback_data.error_message or "处理失败"
        )
    
    return {"success": True}
