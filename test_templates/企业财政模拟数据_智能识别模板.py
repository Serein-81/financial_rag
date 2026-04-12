"""
创建最智能的综合测试模板
包含完整的企业财政模拟数据，支持所有列名识别格式
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def create_intelligent_template():
    """创建最智能的综合测试模板"""
    wb = Workbook()

    # Sheet 1: 使用说明
    ws_info = wb.active
    ws_info.title = '使用说明'

    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    info_data = [
        ['═' * 60],
        ['企业财政智能识别测试数据'],
        ['═' * 60],
        [''],
        ['【功能说明】'],
        ['本文件为企业财政智能识别系统的测试数据模板'],
        ['系统支持自动识别各种格式的Excel列名，无需手动匹配'],
        [''],
        ['【数据说明】'],
        ['包含3家企业的2022-2024年度财务数据，共9条记录'],
        ['涵盖各种财务指标：收入、支出、税费、工资等'],
        [''],
        ['【列名识别示例】'],
        ['系统可识别的列名格式（部分示例）：'],
        ['• 总收入 / Revenue / Sales / 收入 / 营业额'],
        ['• 应税销售额 / Taxable Sales / 销售 / Taxable Revenue'],
        ['• 进项税额 / Input Tax / VAT In / 收票税额'],
        ['• 销项税额 / Output Tax / VAT Out / 开票税额'],
        [''],
        ['【使用方法】'],
        ['1. 下载本模板文件'],
        ['2. 在"财务数据"工作表中查看示例数据'],
        ['3. 修改为企业自己的财务数据（或直接上传测试）'],
        ['4. 在系统中上传本文件进行测试'],
        [''],
        ['【识别结果】'],
        ['本模板使用标准列名，识别率100%，支持全部21个字段'],
        [''],
        ['【注意事项】'],
        ['• 年份列必须填写（2000-2100）'],
        ['• 金额列建议填写数字，如：1000000'],
        ['• 税率列建议填写小数，如：0.13表示13%'],
        ['• 支持直接上传.xlsx和.xls格式'],
    ]

    for row in info_data:
        ws_info.append(row)

    ws_info.column_dimensions['A'].width = 65

    for row in ws_info.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and cell.value.startswith('═'):
                cell.font = Font(bold=True, size=12, color="366092")

    for row in ws_info.iter_rows(min_row=6, max_row=15):
        for cell in row:
            if cell.value and cell.value.startswith('【'):
                cell.font = Font(bold=True, color="2196F3")

    # Sheet 2: 财务数据
    ws_data = wb.create_sheet('财务数据')

    headers = [
        '财务年度', '周期类型', '周期开始日期', '周期结束日期',
        '总收入', '应税销售额', '免税销售额',
        '总支出', '可抵扣支出', '不可抵扣支出',
        '进项税额', '销项税额',
        '增值税率', '应纳税所得额', '企业所得税率',
        '是否小微企业', '工资薪金', '专项附加扣除',
        '发票总数', '进项发票数', '销项发票数'
    ]

    # 企业财政模拟数据
    # 包含3家企业（A公司、B公司、C公司），每年3条记录，共9条数据
    financial_data = [
        # A公司 - 大型企业（2024）
        [2024, 'yearly', '2024-01-01', '2024-12-31',
         5000000, 4500000, 500000,
         3500000, 2800000, 700000,
         585000, 450000,
         0.13, 1500000, 0.25,
         '否', 1200000, 150000,
         8500, 6200, 2300],

        # A公司 - 大型企业（2023）
        [2023, 'yearly', '2023-01-01', '2023-12-31',
         4800000, 4300000, 500000,
         3300000, 2600000, 700000,
         559000, 430000,
         0.13, 1500000, 0.25,
         '否', 1150000, 140000,
         8200, 6000, 2200],

        # A公司 - 大型企业（2022）
        [2022, 'yearly', '2022-01-01', '2022-12-31',
         4500000, 4000000, 500000,
         3100000, 2400000, 700000,
         520000, 400000,
         0.13, 1400000, 0.25,
         '否', 1100000, 130000,
         7900, 5800, 2100],

        # B公司 - 小微企业（2024）
        [2024, 'yearly', '2024-01-01', '2024-12-31',
         800000, 700000, 100000,
         550000, 450000, 100000,
         91000, 70000,
         0.13, 250000, 0.20,
         '是', 200000, 30000,
         1800, 1350, 450],

        # B公司 - 小微企业（2023）
        [2023, 'yearly', '2023-01-01', '2023-12-31',
         750000, 650000, 100000,
         520000, 420000, 100000,
         84500, 65000,
         0.13, 230000, 0.20,
         '是', 190000, 28000,
         1700, 1280, 420],

        # B公司 - 小微企业（2022）
        [2022, 'yearly', '2022-01-01', '2022-12-31',
         680000, 580000, 100000,
         480000, 380000, 100000,
         75400, 58000,
         0.13, 200000, 0.20,
         '是', 180000, 25000,
         1550, 1180, 370],

        # C公司 - 中型企业（2024）
        [2024, 'yearly', '2024-01-01', '2024-12-31',
         2800000, 2500000, 300000,
         1900000, 1500000, 400000,
         325000, 250000,
         0.13, 900000, 0.25,
         '否', 650000, 85000,
         4500, 3300, 1200],

        # C公司 - 中型企业（2023）
        [2023, 'yearly', '2023-01-01', '2023-12-31',
         2600000, 2300000, 300000,
         1750000, 1350000, 400000,
         299000, 230000,
         0.13, 850000, 0.25,
         '否', 620000, 80000,
         4200, 3100, 1100],

        # C公司 - 中型企业（2022）
        [2022, 'yearly', '2022-01-01', '2022-12-31',
         2400000, 2100000, 300000,
         1600000, 1200000, 400000,
         273000, 210000,
         0.13, 800000, 0.25,
         '否', 580000, 75000,
         3900, 2900, 1000],
    ]

    ws_data.append(headers)

    for row in financial_data:
        ws_data.append(row)

    # 样式设置
    header_style = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style='thin', color='CCCCCC')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for cell in ws_data[1]:
        cell.fill = header_style
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row):
        for cell in row:
            cell.border = border
            if cell.column > 1:
                cell.alignment = Alignment(horizontal='right')
                if isinstance(cell.value, (int, float)) and cell.column in [5, 6, 7, 8, 9, 10, 11, 12, 15, 17, 18]:
                    cell.number_format = '#,##0'

    for column in ws_data.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 3, 20)
        ws_data.column_dimensions[column_letter].width = adjusted_width

    ws_data.row_dimensions[1].height = 25

    output_path = os.path.join(
        os.path.dirname(__file__),
        '企业财政模拟数据_智能识别模板.xlsx'
    )

    wb.save(output_path)
    print(f"✓ 已创建最智能的综合测试模板:")
    print(f"  {output_path}")
    print()
    print("模板特点:")
    print("  • 包含9条真实企业财政模拟数据")
    print("  • 3家企业（A/B/C公司）x 3年（2022-2024）")
    print("  • 100%识别率，支持全部21个字段")
    print("  • 包含详细使用说明")
    print("  • 包含小微企业和大型企业对比数据")

if __name__ == '__main__':
    create_intelligent_template()
