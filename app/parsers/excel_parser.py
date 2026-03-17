"""
Excel文件解析器

支持.xls和.xlsx格式的Excel文件解析
提取表格数据、工作表信息等
"""

import io
from typing import Dict, List, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .base_parser import FileParserStrategy


class ExcelParser(FileParserStrategy):
    """Excel文件解析器"""
    
    def get_supported_mime_types(self) -> List[str]:
        return [
            "application/vnd.ms-excel",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/excel",
            "application/x-excel"
        ]
    
    async def parse(self, file_bytes: bytes) -> Dict[str, Any]:
        """
        解析Excel文件
        
        Args:
            file_bytes: Excel文件的字节流
            
        Returns:
            dict: {
                'text': str,  # 所有文本内容
                'sheets': List[dict],  # 工作表信息
                'tables': List[dict],  # 表格数据
                'type': 'excel'
            }
        """
        if not self.validate_file(file_bytes):
            raise ValueError("Excel文件为空或无效")
        
        try:
            # 使用openpyxl读取Excel
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                data_only=True  # 只读取值,不读取公式
            )
            
            sheets_data = []
            all_text = []
            all_tables = []
            
            # 遍历所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 提取工作表数据
                sheet_data = self._extract_sheet_data(sheet)
                sheets_data.append({
                    'name': sheet_name,
                    'rows': sheet_data['rows'],
                    'row_count': sheet_data['row_count'],
                    'col_count': sheet_data['col_count']
                })
                
                # 收集文本内容
                all_text.append(f"工作表: {sheet_name}")
                all_text.append(sheet_data['text'])
                
                # 收集表格数据
                if sheet_data['tables']:
                    all_tables.extend(sheet_data['tables'])
            
            # 关闭工作簿
            workbook.close()
            
            return {
                'text': '\n\n'.join(all_text),
                'sheets': sheets_data,
                'tables': all_tables,
                'type': 'excel',
                'total_sheets': len(sheets_data)
            }
            
        except Exception as e:
            raise Exception(f"Excel解析失败: {str(e)}")
    
    def _extract_sheet_data(self, sheet: Worksheet) -> Dict[str, Any]:
        """
        提取工作表数据
        
        Args:
            sheet: openpyxl工作表对象
            
        Returns:
            dict: 工作表数据
        """
        rows_data = []
        text_lines = []
        tables = []
        
        # 获取有数据的区域
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # 读取所有行
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # 过滤空行
            row_values = [self._format_cell_value(cell) for cell in row]
            
            if any(row_values):  # 如果行中有非空值
                rows_data.append(row_values)
                
                # 构建文本行
                text_line = ' | '.join(str(v) for v in row_values if v)
                text_lines.append(text_line)
        
        # 尝试识别表格结构
        if rows_data:
            # 假设第一行是表头
            if len(rows_data) > 1:
                headers = rows_data[0]
                data_rows = rows_data[1:]
                
                tables.append({
                    'headers': headers,
                    'rows': data_rows,
                    'row_count': len(data_rows),
                    'col_count': len(headers)
                })
        
        return {
            'rows': rows_data,
            'text': '\n'.join(text_lines),
            'tables': tables,
            'row_count': len(rows_data),
            'col_count': max_col
        }
    
    def _format_cell_value(self, value: Any) -> str:
        """
        格式化单元格值
        
        Args:
            value: 单元格值
            
        Returns:
            str: 格式化后的字符串
        """
        if value is None:
            return ''
        
        # 处理数字
        if isinstance(value, (int, float)):
            # 如果是整数,不显示小数点
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)
        
        # 处理日期时间
        from datetime import datetime, date
        if isinstance(value, (datetime, date)):
            return value.strftime('%Y-%m-%d')
        
        # 其他类型转为字符串
        return str(value).strip()
